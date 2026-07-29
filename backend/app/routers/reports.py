from datetime import datetime, timedelta, date, time as datetime_time
from zoneinfo import ZoneInfo
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
import io
import openpyxl
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models import (
    TankStockLedger, LocationAccountingDaySetting, MaterialBalanceTemplate,
    MaterialBalanceTemplateColumn, OperationTransaction, OperationTransactionValue, OperationType,
    MovementMapping, MovementMappingItem, MovementMappingComparison, User,
    OperationTemplate,
)
from app.schemas import (
    TankStockLedgerResponse, TankStockLedgerSummaryResponse,
    TankStockLedgerDailySummaryResponse, OutTurnReportResponse,
    MaterialBalanceDynamicReportResponse, FSOOTRReportResponse,
    FSOMaterialBalanceReportResponse, FSOOutturnReportResponse,
)
from app.dependencies.auth import get_current_user_from_token
from app.dependencies.permissions import require_user_permission
from app.services.audit_service import create_audit_log
from app.utils.helpers import (
    safe_float, clean_optional_text, get_transaction_ticket_number,
    get_location_by_code, get_current_user_display_name, get_asset_by_code,
    normalize_code,
)
from app.config import APPROVED_TRANSACTION_STATUS
from app.routers.tank_stock_ledger import (
    get_filtered_tank_stock_ledger_rows,
    build_tank_stock_ledger_response,
    build_date_range,
    get_active_location_day_setting,
    compute_accounting_date,
    combine_operation_datetime,
    get_tank_stock_rows_for_daily_summary,
    get_stock_snapshot_values,
    get_ledger_operation_datetime,
    build_out_turn_report_response,
    build_tank_stock_daily_summary_rows,
    get_out_turn_report_rows,
)
from app.services.transaction_helpers import parse_date_filter
from app.services.material_balance_helpers import (
    normalize_material_balance_category,
    normalize_material_balance_code_value,
    get_active_material_balance_template_for_location,
    get_active_material_balance_template_columns,
    get_movement_value_for_unit,
    get_snapshot_value_for_unit,
    should_row_match_material_balance_column,
    get_global_internal_transfer_operation_codes,
    should_row_be_in_book_closing_formula,
    calculate_book_closing_from_eligible_ledger_rows,
)

router = APIRouter(prefix="/reports", tags=["Reports"])


def recompute_mapping_comparison(db: Session, mapping_id: int):
    items = db.query(MovementMappingItem).filter(MovementMappingItem.mapping_id == mapping_id).all()

    source = [i for i in items if str(i.role).upper() == "SOURCE"]
    target = [i for i in items if str(i.role).upper() == "TARGET"]

    source_qty = sum(safe_float(i.qty_bbl) for i in source)
    source_water = sum(safe_float(i.water_bbl) for i in source)
    source_nsv = sum(safe_float(i.nsv_bbl) for i in source)

    target_qty = sum(safe_float(i.qty_bbl) for i in target)
    target_water = sum(safe_float(i.water_bbl) for i in target)
    target_nsv = sum(safe_float(i.nsv_bbl) for i in target)

    diff_nsv = target_nsv - source_nsv
    diff_pct = (diff_nsv / source_nsv * 100) if source_nsv else 0

    summary = {
        "source": {"qty_bbl": source_qty, "water_bbl": source_water, "nsv_bbl": source_nsv},
        "target": {"qty_bbl": target_qty, "water_bbl": target_water, "nsv_bbl": target_nsv},
        "diff": {"nsv_bbl": diff_nsv, "nsv_percent": diff_pct},
    }

    cmp_row = db.query(MovementMappingComparison).filter(MovementMappingComparison.mapping_id == mapping_id).first()
    if not cmp_row:
        cmp_row = MovementMappingComparison(mapping_id=mapping_id)
        db.add(cmp_row)

    cmp_row.source_qty_bbl = source_qty
    cmp_row.source_water_bbl = source_water
    cmp_row.source_nsv_bbl = source_nsv

    cmp_row.target_qty_bbl = target_qty
    cmp_row.target_water_bbl = target_water
    cmp_row.target_nsv_bbl = target_nsv

    cmp_row.diff_nsv_bbl = diff_nsv
    cmp_row.diff_nsv_percent = diff_pct

    cmp_row.summary_json = summary
    cmp_row.updated_at = datetime.now()

    db.flush()
    return cmp_row


def build_mapping_response(db: Session, mapping: MovementMapping):
    items = db.query(MovementMappingItem).filter(MovementMappingItem.mapping_id == mapping.id).all()

    source = [i for i in items if str(i.role).upper() == "SOURCE"]
    target = [i for i in items if str(i.role).upper() == "TARGET"]

    return {
        "id": mapping.id,
        "location_code": mapping.location_code,
        "mapping_status": mapping.mapping_status,
        "mapping_label": mapping.mapping_label,
        "source_tickets": [{"ticket_number": i.ticket_number, "ticket_label": i.ticket_label, "movement_date": i.movement_date, "operation_number": i.operation_number} for i in source],
        "target_tickets": [{"ticket_number": i.ticket_number, "ticket_label": i.ticket_label, "movement_date": i.movement_date, "operation_number": i.operation_number} for i in target],
}


def _build_out_turn_transaction_payload_map(db: Session, transaction_ids: list[int]):
    if not transaction_ids:
        return {}

    rows = (
        db.query(OperationTransactionValue)
        .filter(
            OperationTransactionValue.transaction_id.in_(transaction_ids),
            OperationTransactionValue.field_code.in_([
                "tank_gauging_payload",
                "multi_tank_payload",
                "net_nsv",
            ]),
        )
        .all()
    )

    payload_map = {}

    for row in rows:
        if row.transaction_id not in payload_map:
            payload_map[row.transaction_id] = {}

        payload_map[row.transaction_id][row.field_code] = (
            row.field_value if isinstance(row.field_value, dict) else {}
        )

    return payload_map


def _parse_operation_datetime(operation_date: date, payload: dict):
    inputs = payload.get("inputs") or {}

    gauging_date = clean_optional_text(inputs.get("gaugingDate"))
    gauging_time = clean_optional_text(inputs.get("gaugingTime"))

    if gauging_date and gauging_time:
        try:
            return datetime.fromisoformat(f"{gauging_date}T{gauging_time}")
        except Exception:
            pass

    return datetime.combine(operation_date, datetime_time(0, 0))


def _extract_out_turn_values_from_transaction(db: Session, transaction: OperationTransaction):
    template = (
        db.query(OperationTemplate)
        .filter(OperationTemplate.id == transaction.operation_template_id)
        .first()
    )

    if not template:
        return []

    layout = str(template.entry_layout_type or "").strip()

    payload_map = _build_out_turn_transaction_payload_map(db, [transaction.id])
    payloads = payload_map.get(transaction.id, {})

    if layout == "Tank Gauging":
        return _extract_tank_gauging_out_turn(transaction, payloads.get("tank_gauging_payload") or {})
    elif layout == "Multi-Tank Before/After":
        return _extract_multi_tank_out_turn(transaction, payloads.get("multi_tank_payload") or {})
    elif layout == "Stock Movement":
        return _extract_stock_movement_out_turn(transaction, payloads.get("net_nsv"))
    elif layout == "Vessel Cycle":
        return _extract_stock_movement_out_turn(transaction, payloads.get("net_nsv"))

    return []


def _extract_tank_gauging_out_turn(transaction: OperationTransaction, payload: dict):
    calculated = payload.get("calculated") or {}
    inputs = payload.get("inputs") or {}

    nsv = safe_float(calculated.get("nsvBbl"))
    gsv = safe_float(calculated.get("gsvBbl"))
    lt = safe_float(calculated.get("lt"))
    mt = safe_float(calculated.get("mt"))

    if nsv == 0 and gsv == 0:
        return []

    operation_datetime = _parse_operation_datetime(transaction.operation_date, payload)

    return [
        {
            "transaction_id": transaction.id,
            "ticket_number": get_transaction_ticket_number(transaction),
            "operation_number": transaction.operation_number,
            "accounting_date": transaction.operation_date,
            "operation_datetime": operation_datetime,
            "location_code": transaction.origin_location_code,
            "tank_asset_code": transaction.primary_asset_code,
            "product_name": transaction.product_name,
            "tank_operation_code": clean_optional_text(inputs.get("tankOperationCode")) or "",
            "tank_operation_label": clean_optional_text(inputs.get("tankOperationLabel")) or "",
            "tank_operation_category": clean_optional_text(inputs.get("tankOperationCategory")) or "",
            "tank_operation_sign": clean_optional_text(inputs.get("tankOperationSign")) or "",
            "gsv": gsv,
            "nsv": nsv,
            "lt": lt,
            "mt": mt,
            "status": transaction.status,
        }
    ]


def _extract_multi_tank_out_turn(transaction: OperationTransaction, payload: dict):
    per_tank_after = ((payload.get("perTank") or {}).get("after") or {})
    inputs = payload.get("inputs") or {}

    if not per_tank_after:
        return []

    operation_datetime = _parse_operation_datetime(transaction.operation_date, payload)

    results = []

    for tank_key, tank_data in per_tank_after.items():
        if not isinstance(tank_data, dict):
            continue

        nsv = safe_float(tank_data.get("NSV"))
        gsv = safe_float(tank_data.get("GSV"))
        lt = safe_float(tank_data.get("LT"))
        mt = safe_float(tank_data.get("MT"))

        if nsv == 0 and gsv == 0:
            continue

        results.append(
            {
                "transaction_id": transaction.id,
                "ticket_number": get_transaction_ticket_number(transaction),
                "operation_number": transaction.operation_number,
                "accounting_date": transaction.operation_date,
                "operation_datetime": operation_datetime,
                "location_code": transaction.origin_location_code,
                "tank_asset_code": tank_key,
                "product_name": transaction.product_name,
                "tank_operation_code": clean_optional_text(inputs.get("tankOperationCode")) or "",
                "tank_operation_label": clean_optional_text(inputs.get("tankOperationLabel")) or "",
                "tank_operation_category": clean_optional_text(inputs.get("tankOperationCategory")) or "",
                "tank_operation_sign": clean_optional_text(inputs.get("tankOperationSign")) or "",
                "gsv": gsv,
                "nsv": nsv,
                "lt": lt,
                "mt": mt,
                "status": transaction.status,
            }
        )

    return results


def _extract_stock_movement_out_turn(transaction: OperationTransaction, net_nsv_value):
    nsv = safe_float(net_nsv_value)

    if nsv == 0:
        return []

    operation_datetime = transaction.operation_start_datetime
    if operation_datetime is None:
        operation_datetime = datetime.combine(transaction.operation_date, datetime_time(0, 0))

    return [
        {
            "transaction_id": transaction.id,
            "ticket_number": get_transaction_ticket_number(transaction),
            "operation_number": transaction.operation_number,
            "accounting_date": transaction.operation_date,
            "operation_datetime": operation_datetime,
            "location_code": transaction.origin_location_code,
            "tank_asset_code": transaction.primary_asset_code,
            "product_name": transaction.product_name,
            "tank_operation_code": "",
            "tank_operation_label": "",
            "tank_operation_category": "",
            "tank_operation_sign": "",
            "gsv": 0,
            "nsv": nsv,
            "lt": 0,
            "mt": 0,
            "status": transaction.status,
        }
    ]


def get_out_turn_report_rows_from_transactions(
    db: Session,
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
):
    query = (
        db.query(OperationTransaction)
        .join(OperationTemplate, OperationTransaction.operation_template_id == OperationTemplate.id)
        .join(OperationType, OperationTemplate.operation_type_code == OperationType.operation_type_code)
        .filter(
            OperationTransaction.status == "Approved",
            OperationType.applicable_asset_type_code == "TANK",
            OperationTemplate.entry_layout_type.in_([
                "Tank Gauging",
                "Multi-Tank Before/After",
                "Stock Movement",
                "Vessel Cycle",
            ]),
        )
    )

    cleaned_location_code = clean_optional_text(location_code)
    cleaned_tank_asset_code = clean_optional_text(tank_asset_code)
    cleaned_product_name = clean_optional_text(product_name)

    if cleaned_location_code:
        query = query.filter(
            OperationTransaction.origin_location_code.ilike(cleaned_location_code)
        )

    if cleaned_tank_asset_code:
        query = query.filter(
            OperationTransaction.primary_asset_code.ilike(cleaned_tank_asset_code)
        )

    if cleaned_product_name:
        query = query.filter(
            OperationTransaction.product_name.ilike(cleaned_product_name)
        )

    date_from_value = parse_date_filter(date_from, "Date From")
    date_to_value = parse_date_filter(date_to, "Date To")

    if date_from_value:
        query = query.filter(OperationTransaction.operation_date >= date_from_value)

    if date_to_value:
        query = query.filter(OperationTransaction.operation_date <= date_to_value)

    transactions = query.order_by(
        OperationTransaction.operation_date.asc(),
        OperationTransaction.id.asc(),
    ).all()

    tx_ids = [tx.id for tx in transactions]
    payload_map = _build_out_turn_transaction_payload_map(db, tx_ids)

    template_cache = {}

    extracted_rows = []

    for tx in transactions:
        if tx.operation_template_id is None:
            continue

        if tx.operation_template_id not in template_cache:
            template_cache[tx.operation_template_id] = (
                db.query(OperationTemplate)
                .filter(OperationTemplate.id == tx.operation_template_id)
                .first()
            )

        template = template_cache[tx.operation_template_id]

        if not template:
            continue

        payloads = payload_map.get(tx.id, {})
        entry_layout = str(template.entry_layout_type or "").strip()

        if entry_layout == "Tank Gauging":
            extracted_rows.extend(
                _extract_tank_gauging_out_turn(tx, payloads.get("tank_gauging_payload") or {})
            )
        elif entry_layout == "Multi-Tank Before/After":
            extracted_rows.extend(
                _extract_multi_tank_out_turn(tx, payloads.get("multi_tank_payload") or {})
            )
        elif entry_layout in ("Stock Movement", "Vessel Cycle"):
            extracted_rows.extend(
                _extract_stock_movement_out_turn(tx, payloads.get("net_nsv"))
            )

    grouped_rows = {}

    for row in extracted_rows:
        key = (
            row["location_code"],
            row["tank_asset_code"],
        )

        if key not in grouped_rows:
            grouped_rows[key] = []

        grouped_rows[key].append(row)

    per_tank_rows = []

    for key, group_rows in grouped_rows.items():
        sorted_group = sorted(
            group_rows,
            key=lambda item: (
                item["accounting_date"] or date.min,
                item["operation_datetime"] or datetime.min,
                item["transaction_id"],
            ),
        )

        previous_gsv = 0
        previous_nsv = 0
        previous_lt = 0
        previous_mt = 0
        sequence = 0

        for item in sorted_group:
            current_gsv = safe_float(item["gsv"])
            current_nsv = safe_float(item["nsv"])
            current_lt = safe_float(item["lt"])
            current_mt = safe_float(item["mt"])

            per_tank_rows.append(
                {
                    "sequence": sequence,
                    "location_code": item["location_code"],
                    "tank_asset_code": item["tank_asset_code"],
                    "product_name": item["product_name"],
                    "transaction_id": item["transaction_id"],
                    "ticket_number": item["ticket_number"],
                    "operation_number": item["operation_number"],
                    "accounting_date": item["accounting_date"],
                    "operation_datetime": item["operation_datetime"],
                    "tank_operation_code": item["tank_operation_code"],
                    "tank_operation_label": item["tank_operation_label"],
                    "tank_operation_category": item["tank_operation_category"],
                    "tank_operation_sign": item["tank_operation_sign"],
                    "previous_gsv": previous_gsv,
                    "previous_nsv": previous_nsv,
                    "previous_lt": previous_lt,
                    "previous_mt": previous_mt,
                    "stock_after_gsv": current_gsv,
                    "stock_after_nsv": current_nsv,
                    "stock_after_lt": current_lt,
                    "stock_after_mt": current_mt,
                    "signed_net_gsv": current_gsv - previous_gsv,
                    "signed_net_nsv": current_nsv - previous_nsv,
                    "signed_net_lt": current_lt - previous_lt,
                    "signed_net_mt": current_mt - previous_mt,
                    "status": item["status"],
                }
            )

            previous_gsv = current_gsv
            previous_nsv = current_nsv
            previous_lt = current_lt
            previous_mt = current_mt
            sequence += 1

    per_tank_rows.sort(
        key=lambda r: (
            r["accounting_date"] or date.min,
            r["operation_datetime"] or datetime.min,
            r["location_code"] or "",
            r["tank_asset_code"] or "",
            r["transaction_id"] or 0,
        )
    )

    return per_tank_rows


def build_out_turn_report_response_from_transaction(row: dict, db: Session):
    from app.services.stock_ledger_helpers import build_out_turn_report_response_from_values

    return build_out_turn_report_response_from_values(
        params={
            "ledger_id": row.get("transaction_id", 0) * 1000000 + (row.get("sequence") or 0),
            "transaction_id": row.get("transaction_id"),
            "ticket_number": row.get("ticket_number") or "",
            "operation_number": row.get("operation_number") or "",
            "accounting_date": row.get("accounting_date"),
            "operation_datetime": row.get("operation_datetime"),
            "location_code": row.get("location_code") or "",
            "tank_asset_code": row.get("tank_asset_code") or "",
            "tank_asset_name": getattr(get_asset_by_code(row.get("tank_asset_code") or "", db), "asset_name", ""),
            "product_name": row.get("product_name") or "",
            "tank_operation_code": row.get("tank_operation_code") or "",
            "tank_operation_label": row.get("tank_operation_label") or "",
            "tank_operation_category": row.get("tank_operation_category") or "",
            "tank_operation_sign": row.get("tank_operation_sign") or "",
            "previous_gsv": row.get("previous_gsv") or 0,
            "previous_nsv": row.get("previous_nsv") or 0,
            "previous_lt": row.get("previous_lt") or 0,
            "previous_mt": row.get("previous_mt") or 0,
            "stock_after_gsv": row.get("stock_after_gsv") or 0,
            "stock_after_nsv": row.get("stock_after_nsv") or 0,
            "stock_after_lt": row.get("stock_after_lt") or 0,
            "stock_after_mt": row.get("stock_after_mt") or 0,
            "signed_net_gsv": row.get("signed_net_gsv"),
            "signed_net_nsv": row.get("signed_net_nsv"),
            "signed_net_lt": row.get("signed_net_lt"),
            "signed_net_mt": row.get("signed_net_mt"),
            "status": "Approved",
            "remarks": None,
        },
        db=db,
    )


def add_volume_values(target: dict, prefix: str, row: TankStockLedger):
    from app.services.stock_ledger_helpers import add_volume_values_from_values
    add_volume_values_from_values(
        target=target,
        prefix=prefix,
        movement_gsv=row.movement_gsv_bbl,
        movement_nsv=row.movement_nsv_bbl,
        movement_lt=row.movement_lt,
        movement_mt=row.movement_mt,
    )


def get_material_balance_rows_for_continuity(
    db: Session,
    location_code: str | None,
    tank_asset_code: str | None,
    product_name: str | None,
    date_to_value: date,
    status: str | None = "Active",
):
    query = db.query(TankStockLedger)

    cleaned_status = clean_optional_text(status)

    if cleaned_status:
        query = query.filter(TankStockLedger.status == cleaned_status)

    query = query.filter(
        TankStockLedger.accounting_date != None,
        TankStockLedger.accounting_date <= date_to_value,
    )

    cleaned_location_code = clean_optional_text(location_code)
    cleaned_tank_asset_code = clean_optional_text(tank_asset_code)
    cleaned_product_name = clean_optional_text(product_name)

    if cleaned_location_code:
        query = query.filter(
            TankStockLedger.location_code.ilike(cleaned_location_code)
        )

    if cleaned_tank_asset_code:
        query = query.filter(
            TankStockLedger.tank_asset_code.ilike(cleaned_tank_asset_code)
        )

    if cleaned_product_name:
        query = query.filter(
            TankStockLedger.product_name.ilike(cleaned_product_name)
        )

    rows = query.all()

    return sorted(
        rows,
        key=lambda row: (
            row.location_code or "",
            row.tank_asset_code or "",
            row.product_name or "",
            row.accounting_date or date.min,
            get_ledger_operation_datetime(row) or datetime.min,
            row.id,
        ),
    )


def build_material_balance_report_rows(
    db: Session,
    ledger_rows: list[TankStockLedger],
    date_from_value: date,
    date_to_value: date,
):
    date_range = build_date_range(date_from_value, date_to_value)

    grouped_rows = {}

    for row in ledger_rows:
        key = (
            row.location_code,
            row.tank_asset_code,
            row.product_name or "",
        )

        if key not in grouped_rows:
            grouped_rows[key] = []

        grouped_rows[key].append(row)

    material_balance_rows = []

    for key, rows in grouped_rows.items():
        location_code, tank_asset_code, product_name_value = key

        location = get_location_by_code(location_code, db)

        sorted_rows = sorted(
            rows,
            key=lambda row: (
                row.accounting_date or date.min,
                get_ledger_operation_datetime(row) or datetime.min,
                row.id,
            ),
        )

        tank_asset_name = ""
        if sorted_rows:
            tank_asset_name = sorted_rows[-1].tank_asset_name or ""

        previous_closing_gsv = 0
        previous_closing_nsv = 0
        previous_closing_lt = 0
        previous_closing_mt = 0

        rows_before_period = [
            row
            for row in sorted_rows
            if row.accounting_date is not None
            and row.accounting_date < date_from_value
        ]

        if rows_before_period:
            last_before_period = rows_before_period[-1]
            previous_snapshot = get_stock_snapshot_values(last_before_period)

            previous_closing_gsv = previous_snapshot["gsv"]
            previous_closing_nsv = previous_snapshot["nsv"]
            previous_closing_lt = previous_snapshot["lt"]
            previous_closing_mt = previous_snapshot["mt"]

        for accounting_date_value in date_range:
            day_rows = [
                row
                for row in sorted_rows
                if row.accounting_date == accounting_date_value
            ]

            day_rows = sorted(
                day_rows,
                key=lambda row: (
                    get_ledger_operation_datetime(row) or datetime.min,
                    row.id,
                ),
            )

            opening_gsv = previous_closing_gsv
            opening_nsv = previous_closing_nsv
            opening_lt = previous_closing_lt
            opening_mt = previous_closing_mt

            opening_rows = [
                row
                for row in day_rows
                if normalize_material_balance_category(
                    row.tank_operation_category
                )
                == "OPENING"
            ]

            if opening_rows:
                opening_snapshot = get_stock_snapshot_values(opening_rows[-1])

                opening_gsv = opening_snapshot["gsv"]
                opening_nsv = opening_snapshot["nsv"]
                opening_lt = opening_snapshot["lt"]
                opening_mt = opening_snapshot["mt"]

            buckets = {
                "receipt_gsv": 0,
                "receipt_nsv": 0,
                "receipt_lt": 0,
                "receipt_mt": 0,
                "production_gsv": 0,
                "production_nsv": 0,
                "production_lt": 0,
                "production_mt": 0,
                "dispatch_gsv": 0,
                "dispatch_nsv": 0,
                "dispatch_lt": 0,
                "dispatch_mt": 0,
                "draining_gsv": 0,
                "draining_nsv": 0,
                "draining_lt": 0,
                "draining_mt": 0,
                "other_in_gsv": 0,
                "other_in_nsv": 0,
                "other_in_lt": 0,
                "other_in_mt": 0,
                "other_out_gsv": 0,
                "other_out_nsv": 0,
                "other_out_lt": 0,
                "other_out_mt": 0,
            }

            for row in day_rows:
                sign = str(row.tank_operation_sign or "").upper()
                category = normalize_material_balance_category(
                    row.tank_operation_category
                )

                if sign == "IN":
                    if category == "RECEIPT":
                        add_volume_values(buckets, "receipt", row)
                    elif category == "PRODUCTION":
                        add_volume_values(buckets, "production", row)
                    else:
                        add_volume_values(buckets, "other_in", row)

                elif sign == "OUT":
                    if category == "DISPATCH":
                        add_volume_values(buckets, "dispatch", row)
                    elif category == "DRAINING":
                        add_volume_values(buckets, "draining", row)
                    else:
                        add_volume_values(buckets, "other_out", row)

            total_in_gsv = (
                buckets["receipt_gsv"]
                + buckets["production_gsv"]
                + buckets["other_in_gsv"]
            )
            total_in_nsv = (
                buckets["receipt_nsv"]
                + buckets["production_nsv"]
                + buckets["other_in_nsv"]
            )
            total_in_lt = (
                buckets["receipt_lt"]
                + buckets["production_lt"]
                + buckets["other_in_lt"]
            )
            total_in_mt = (
                buckets["receipt_mt"]
                + buckets["production_mt"]
                + buckets["other_in_mt"]
            )

            total_out_gsv = (
                buckets["dispatch_gsv"]
                + buckets["draining_gsv"]
                + buckets["other_out_gsv"]
            )
            total_out_nsv = (
                buckets["dispatch_nsv"]
                + buckets["draining_nsv"]
                + buckets["other_out_nsv"]
            )
            total_out_lt = (
                buckets["dispatch_lt"]
                + buckets["draining_lt"]
                + buckets["other_out_lt"]
            )
            total_out_mt = (
                buckets["dispatch_mt"]
                + buckets["draining_mt"]
                + buckets["other_out_mt"]
            )

            book_closing_gsv = opening_gsv + total_in_gsv - total_out_gsv
            book_closing_nsv = opening_nsv + total_in_nsv - total_out_nsv
            book_closing_lt = opening_lt + total_in_lt - total_out_lt
            book_closing_mt = opening_mt + total_in_mt - total_out_mt

            actual_closing_gsv = book_closing_gsv
            actual_closing_nsv = book_closing_nsv
            actual_closing_lt = book_closing_lt
            actual_closing_mt = book_closing_mt

            last_ticket_number = None

            if day_rows:
                closing_rows = [
                    row
                    for row in day_rows
                    if normalize_material_balance_category(
                        row.tank_operation_category
                    )
                    == "CLOSING"
                ]

                if closing_rows:
                    closing_source_row = closing_rows[-1]
                else:
                    closing_source_row = day_rows[-1]

                closing_snapshot = get_stock_snapshot_values(closing_source_row)

                actual_closing_gsv = closing_snapshot["gsv"]
                actual_closing_nsv = closing_snapshot["nsv"]
                actual_closing_lt = closing_snapshot["lt"]
                actual_closing_mt = closing_snapshot["mt"]

                last_ticket_number = closing_source_row.ticket_number

            else:
                actual_closing_gsv = opening_gsv
                actual_closing_nsv = opening_nsv
                actual_closing_lt = opening_lt
                actual_closing_mt = opening_mt

                book_closing_gsv = opening_gsv
                book_closing_nsv = opening_nsv
                book_closing_lt = opening_lt
                book_closing_mt = opening_mt

            loss_gain_gsv = actual_closing_gsv - book_closing_gsv
            loss_gain_nsv = actual_closing_nsv - book_closing_nsv
            loss_gain_lt = actual_closing_lt - book_closing_lt
            loss_gain_mt = actual_closing_mt - book_closing_mt

            material_balance_rows.append(
                {
                    "accounting_date": accounting_date_value,
                    "location_code": location_code,
                    "location_name": location.location_name if location else "",
                    "tank_asset_code": tank_asset_code,
                    "tank_asset_name": tank_asset_name,
                    "product_name": product_name_value or None,
                    "opening_gsv_bbl": round(opening_gsv, 3),
                    "opening_nsv_bbl": round(opening_nsv, 3),
                    "opening_lt": round(opening_lt, 3),
                    "opening_mt": round(opening_mt, 3),
                    "receipt_gsv_bbl": round(buckets["receipt_gsv"], 3),
                    "receipt_nsv_bbl": round(buckets["receipt_nsv"], 3),
                    "receipt_lt": round(buckets["receipt_lt"], 3),
                    "receipt_mt": round(buckets["receipt_mt"], 3),
                    "production_gsv_bbl": round(buckets["production_gsv"], 3),
                    "production_nsv_bbl": round(buckets["production_nsv"], 3),
                    "production_lt": round(buckets["production_lt"], 3),
                    "production_mt": round(buckets["production_mt"], 3),
                    "dispatch_gsv_bbl": round(buckets["dispatch_gsv"], 3),
                    "dispatch_nsv_bbl": round(buckets["dispatch_nsv"], 3),
                    "dispatch_lt": round(buckets["dispatch_lt"], 3),
                    "dispatch_mt": round(buckets["dispatch_mt"], 3),
                    "draining_gsv_bbl": round(buckets["draining_gsv"], 3),
                    "draining_nsv_bbl": round(buckets["draining_nsv"], 3),
                    "draining_lt": round(buckets["draining_lt"], 3),
                    "draining_mt": round(buckets["draining_mt"], 3),
                    "other_in_gsv_bbl": round(buckets["other_in_gsv"], 3),
                    "other_in_nsv_bbl": round(buckets["other_in_nsv"], 3),
                    "other_in_lt": round(buckets["other_in_lt"], 3),
                    "other_in_mt": round(buckets["other_in_mt"], 3),
                    "other_out_gsv_bbl": round(buckets["other_out_gsv"], 3),
                    "other_out_nsv_bbl": round(buckets["other_out_nsv"], 3),
                    "other_out_lt": round(buckets["other_out_lt"], 3),
                    "other_out_mt": round(buckets["other_out_mt"], 3),
                    "total_in_gsv_bbl": round(total_in_gsv, 3),
                    "total_in_nsv_bbl": round(total_in_nsv, 3),
                    "total_in_lt": round(total_in_lt, 3),
                    "total_in_mt": round(total_in_mt, 3),
                    "total_out_gsv_bbl": round(total_out_gsv, 3),
                    "total_out_nsv_bbl": round(total_out_nsv, 3),
                    "total_out_lt": round(total_out_lt, 3),
                    "total_out_mt": round(total_out_mt, 3),
                    "book_closing_gsv_bbl": round(book_closing_gsv, 3),
                    "book_closing_nsv_bbl": round(book_closing_nsv, 3),
                    "book_closing_lt": round(book_closing_lt, 3),
                    "book_closing_mt": round(book_closing_mt, 3),
                    "actual_closing_gsv_bbl": round(actual_closing_gsv, 3),
                    "actual_closing_nsv_bbl": round(actual_closing_nsv, 3),
                    "actual_closing_lt": round(actual_closing_lt, 3),
                    "actual_closing_mt": round(actual_closing_mt, 3),
                    "loss_gain_gsv_bbl": round(loss_gain_gsv, 3),
                    "loss_gain_nsv_bbl": round(loss_gain_nsv, 3),
                    "loss_gain_lt": round(loss_gain_lt, 3),
                    "loss_gain_mt": round(loss_gain_mt, 3),
                    "rows_count": len(day_rows),
                    "last_ticket_number": last_ticket_number,
                }
            )

            previous_closing_gsv = actual_closing_gsv
            previous_closing_nsv = actual_closing_nsv
            previous_closing_lt = actual_closing_lt
            previous_closing_mt = actual_closing_mt

    return sorted(
        material_balance_rows,
        key=lambda row: (
            row["accounting_date"],
            row["location_code"],
            row["tank_asset_code"] or "",
            row["product_name"] or "",
        ),
    )


def consolidate_material_balance_rows_by_location(
    tank_wise_rows: list[dict],
):
    consolidated_map = {}

    for row in tank_wise_rows:
        key = (
            row["accounting_date"],
            row["location_code"],
            row["product_name"] or "",
        )

        if key not in consolidated_map:
            consolidated_map[key] = {
                "accounting_date": row["accounting_date"],
                "location_code": row["location_code"],
                "location_name": row["location_name"],
                "tank_asset_code": None,
                "tank_asset_name": "All Tanks",
                "product_name": row["product_name"],
                "opening_gsv_bbl": 0,
                "opening_nsv_bbl": 0,
                "opening_lt": 0,
                "opening_mt": 0,
                "receipt_gsv_bbl": 0,
                "receipt_nsv_bbl": 0,
                "receipt_lt": 0,
                "receipt_mt": 0,
                "production_gsv_bbl": 0,
                "production_nsv_bbl": 0,
                "production_lt": 0,
                "production_mt": 0,
                "dispatch_gsv_bbl": 0,
                "dispatch_nsv_bbl": 0,
                "dispatch_lt": 0,
                "dispatch_mt": 0,
                "draining_gsv_bbl": 0,
                "draining_nsv_bbl": 0,
                "draining_lt": 0,
                "draining_mt": 0,
                "other_in_gsv_bbl": 0,
                "other_in_nsv_bbl": 0,
                "other_in_lt": 0,
                "other_in_mt": 0,
                "other_out_gsv_bbl": 0,
                "other_out_nsv_bbl": 0,
                "other_out_lt": 0,
                "other_out_mt": 0,
                "total_in_gsv_bbl": 0,
                "total_in_nsv_bbl": 0,
                "total_in_lt": 0,
                "total_in_mt": 0,
                "total_out_gsv_bbl": 0,
                "total_out_nsv_bbl": 0,
                "total_out_lt": 0,
                "total_out_mt": 0,
                "book_closing_gsv_bbl": 0,
                "book_closing_nsv_bbl": 0,
                "book_closing_lt": 0,
                "book_closing_mt": 0,
                "actual_closing_gsv_bbl": 0,
                "actual_closing_nsv_bbl": 0,
                "actual_closing_lt": 0,
                "actual_closing_mt": 0,
                "loss_gain_gsv_bbl": 0,
                "loss_gain_nsv_bbl": 0,
                "loss_gain_lt": 0,
                "loss_gain_mt": 0,
                "rows_count": 0,
                "last_ticket_number": None,
            }

        target = consolidated_map[key]

        numeric_fields = [
            "opening_gsv_bbl",
            "opening_nsv_bbl",
            "opening_lt",
            "opening_mt",
            "receipt_gsv_bbl",
            "receipt_nsv_bbl",
            "receipt_lt",
            "receipt_mt",
            "production_gsv_bbl",
            "production_nsv_bbl",
            "production_lt",
            "production_mt",
            "dispatch_gsv_bbl",
            "dispatch_nsv_bbl",
            "dispatch_lt",
            "dispatch_mt",
            "draining_gsv_bbl",
            "draining_nsv_bbl",
            "draining_lt",
            "draining_mt",
            "other_in_gsv_bbl",
            "other_in_nsv_bbl",
            "other_in_lt",
            "other_in_mt",
            "other_out_gsv_bbl",
            "other_out_nsv_bbl",
            "other_out_lt",
            "other_out_mt",
            "total_in_gsv_bbl",
            "total_in_nsv_bbl",
            "total_in_lt",
            "total_in_mt",
            "total_out_gsv_bbl",
            "total_out_nsv_bbl",
            "total_out_lt",
            "total_out_mt",
            "book_closing_gsv_bbl",
            "book_closing_nsv_bbl",
            "book_closing_lt",
            "book_closing_mt",
            "actual_closing_gsv_bbl",
            "actual_closing_nsv_bbl",
            "actual_closing_lt",
            "actual_closing_mt",
            "loss_gain_gsv_bbl",
            "loss_gain_nsv_bbl",
            "loss_gain_lt",
            "loss_gain_mt",
        ]

        for field in numeric_fields:
            target[field] += safe_float(row.get(field))

        target["rows_count"] += int(row.get("rows_count") or 0)

        if row.get("last_ticket_number"):
            target["last_ticket_number"] = row.get("last_ticket_number")

    consolidated_rows = []

    for row in consolidated_map.values():
        for key, value in list(row.items()):
            if isinstance(value, float):
                row[key] = round(value, 3)

        consolidated_rows.append(row)

    return sorted(
        consolidated_rows,
        key=lambda row: (
            row["accounting_date"],
            row["location_code"],
            row["product_name"] or "",
        ),
    )


def build_dynamic_material_balance_columns_response(
    columns: list[MaterialBalanceTemplateColumn],
):
    return [
        {
            "column_key": column.column_key,
            "column_label": column.column_label,
            "column_order": column.column_order,
            "column_type": column.column_type,
            "movement_direction": column.movement_direction,
            "include_in_material_balance": column.include_in_material_balance,
            "include_in_book_closing": column.include_in_book_closing,
            "is_internal_transfer": column.is_internal_transfer,
        }
        for column in columns
    ]


def build_dynamic_material_balance_tank_rows(
    db: Session,
    ledger_rows: list[TankStockLedger],
    columns: list[MaterialBalanceTemplateColumn],
    date_from_value: date,
    date_to_value: date,
    unit_key: str = "nsv",
):
    date_range = build_date_range(date_from_value, date_to_value)

    grouped_rows = {}

    for row in ledger_rows:
        key = (
            row.location_code,
            row.tank_asset_code,
            row.product_name or "",
        )

        if key not in grouped_rows:
            grouped_rows[key] = []

        grouped_rows[key].append(row)

    report_rows = []

    for key, rows in grouped_rows.items():
        location_code, tank_asset_code, product_name_value = key
        location = get_location_by_code(location_code, db)

        sorted_rows = sorted(
            rows,
            key=lambda row: (
                row.accounting_date or date.min,
                get_ledger_operation_datetime(row) or datetime.min,
                row.id,
            ),
        )

        tank_asset_name = ""

        if sorted_rows:
            tank_asset_name = sorted_rows[-1].tank_asset_name or ""

        previous_closing_snapshot = {
            "gsv": 0,
            "nsv": 0,
            "lt": 0,
            "mt": 0,
        }

        rows_before_period = [
            row
            for row in sorted_rows
            if row.accounting_date is not None
            and row.accounting_date < date_from_value
        ]

        if rows_before_period:
            previous_closing_snapshot = get_stock_snapshot_values(
                rows_before_period[-1]
            )

        for accounting_date_value in date_range:
            day_rows = [
                row
                for row in sorted_rows
                if row.accounting_date == accounting_date_value
            ]

            day_rows = sorted(
                day_rows,
                key=lambda row: (
                    get_ledger_operation_datetime(row) or datetime.min,
                    row.id,
                ),
            )

            opening_value = get_snapshot_value_for_unit(
                previous_closing_snapshot,
                unit_key,
            )

            explicit_opening_rows = [
                row
                for row in day_rows
                if normalize_material_balance_code_value(
                    row.tank_operation_category
                )
                == "OPENING"
            ]

            if explicit_opening_rows:
                opening_snapshot = get_stock_snapshot_values(
                    explicit_opening_rows[-1]
                )
                opening_value = get_snapshot_value_for_unit(
                    opening_snapshot,
                    unit_key,
                )

            values = {}

            book_closing_value = opening_value
            actual_closing_value = opening_value
            last_ticket_number = None

            for column in columns:
                column_key = column.column_key
                column_type = normalize_material_balance_code_value(
                    column.column_type
                )

                if column_type == "OPENING":
                    values[column_key] = round(opening_value, 3)
                    continue

                if column_type == "MOVEMENT":
                    movement_total = 0

                    for row in day_rows:
                        if should_row_match_material_balance_column(row, column):
                            movement_total += get_movement_value_for_unit(
                                row,
                                unit_key,
                            )

                    values[column_key] = round(movement_total, 3)
                    continue

                if column_type in ["INFO", "FORMULA"]:
                    values[column_key] = 0
                    continue

            book_closing_calculation = calculate_book_closing_from_eligible_ledger_rows(
                opening_value=opening_value,
                day_rows=day_rows,
                columns=columns,
                unit_key=unit_key,
            )

            book_closing_value = book_closing_calculation["book_closing_value"]

            if day_rows:
                explicit_closing_rows = [
                    row
                    for row in day_rows
                    if normalize_material_balance_code_value(
                        row.tank_operation_category
                    )
                    == "CLOSING"
                ]

                if explicit_closing_rows:
                    closing_source_row = explicit_closing_rows[-1]
                else:
                    closing_source_row = day_rows[-1]

                actual_closing_snapshot = get_stock_snapshot_values(
                    closing_source_row
                )

                actual_closing_value = get_snapshot_value_for_unit(
                    actual_closing_snapshot,
                    unit_key,
                )

                last_ticket_number = closing_source_row.ticket_number
            else:
                actual_closing_snapshot = previous_closing_snapshot
                actual_closing_value = opening_value

            loss_gain_value = actual_closing_value - book_closing_value

            for column in columns:
                column_key = column.column_key
                column_type = normalize_material_balance_code_value(
                    column.column_type
                )

                if column_type == "BOOK_CLOSING":
                    values[column_key] = round(book_closing_value, 3)

                elif column_type == "ACTUAL_CLOSING":
                    values[column_key] = round(actual_closing_value, 3)

                elif column_type == "LOSS_GAIN":
                    values[column_key] = round(loss_gain_value, 3)

            report_rows.append(
                {
                    "accounting_date": accounting_date_value,
                    "location_code": location_code,
                    "location_name": location.location_name if location else "",
                    "tank_asset_code": tank_asset_code,
                    "tank_asset_name": tank_asset_name,
                    "product_name": product_name_value or None,
                    "values": values,
                    "rows_count": len(day_rows),
                    "last_ticket_number": last_ticket_number,
                }
            )

            previous_closing_snapshot = {
                "gsv": actual_closing_snapshot.get("gsv", actual_closing_value),
                "nsv": actual_closing_snapshot.get("nsv", actual_closing_value),
                "lt": actual_closing_snapshot.get("lt", 0),
                "mt": actual_closing_snapshot.get("mt", 0),
            }

    return sorted(
        report_rows,
        key=lambda row: (
            row["accounting_date"],
            row["location_code"],
            row["tank_asset_code"] or "",
            row["product_name"] or "",
        ),
    )


def consolidate_dynamic_material_balance_rows_by_location(
    tank_rows: list[dict],
    columns: list[MaterialBalanceTemplateColumn],
):
    consolidated_map = {}

    for row in tank_rows:
        key = (
            row["accounting_date"],
            row["location_code"],
            row["product_name"] or "",
        )

        if key not in consolidated_map:
            consolidated_map[key] = {
                "accounting_date": row["accounting_date"],
                "location_code": row["location_code"],
                "location_name": row["location_name"],
                "tank_asset_code": None,
                "tank_asset_name": "All Tanks",
                "product_name": row["product_name"],
                "values": {},
                "rows_count": 0,
                "last_ticket_number": None,
            }

            for column in columns:
                consolidated_map[key]["values"][column.column_key] = 0

        target = consolidated_map[key]

        for column in columns:
            column_key = column.column_key
            target["values"][column_key] = safe_float(
                target["values"].get(column_key)
            ) + safe_float(row["values"].get(column_key))

        target["rows_count"] += int(row.get("rows_count") or 0)

        if row.get("last_ticket_number"):
            target["last_ticket_number"] = row.get("last_ticket_number")

    consolidated_rows = []

    for row in consolidated_map.values():
        for column in columns:
            column_key = column.column_key
            row["values"][column_key] = round(
                safe_float(row["values"].get(column_key)),
                3,
            )

        consolidated_rows.append(row)

    return sorted(
        consolidated_rows,
        key=lambda row: (
            row["accounting_date"],
            row["location_code"],
            row["product_name"] or "",
        ),
    )


@router.get(
    "/tank-stock-ledger",
    response_model=list[TankStockLedgerResponse],
)
def get_tank_stock_ledger(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = "Active",
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Tank Stock Ledger",
        db,
    )

    ledger_rows = get_filtered_tank_stock_ledger_rows(
        db=db,
        location_code=location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_from=date_from,
        date_to=date_to,
        status=status,
    )

    return [
        build_tank_stock_ledger_response(row, db)
        for row in ledger_rows
    ]


@router.get(
    "/tank-stock-ledger/summary",
    response_model=list[TankStockLedgerSummaryResponse],
)
def get_tank_stock_ledger_summary(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Tank Stock Ledger",
        db,
    )

    ledger_rows = get_filtered_tank_stock_ledger_rows(
        db=db,
        location_code=location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_from=date_from,
        date_to=date_to,
        status="Active",
    )

    summary_map = {}

    for row in ledger_rows:
        key = (
            row.location_code,
            row.tank_asset_code,
            row.product_name or "",
        )

        if key not in summary_map:
            location = get_location_by_code(row.location_code, db)

            summary_map[key] = {
                "location_code": row.location_code,
                "location_name": location.location_name if location else "",
                "tank_asset_code": row.tank_asset_code,
                "tank_asset_name": row.tank_asset_name,
                "product_name": row.product_name,
                "opening_nsv_bbl": 0,
                "total_in_nsv_bbl": 0,
                "total_out_nsv_bbl": 0,
                "closing_nsv_bbl": 0,
                "opening_lt": 0,
                "total_in_lt": 0,
                "total_out_lt": 0,
                "closing_lt": 0,
                "opening_mt": 0,
                "total_in_mt": 0,
                "total_out_mt": 0,
                "closing_mt": 0,
            }

        summary = summary_map[key]

        sign = row.tank_operation_sign
        category = row.tank_operation_category

        movement_nsv = row.movement_nsv_bbl or 0
        movement_lt = row.movement_lt or 0
        movement_mt = row.movement_mt or 0

        if category == "OPENING":
            summary["opening_nsv_bbl"] += movement_nsv
            summary["opening_lt"] += movement_lt
            summary["opening_mt"] += movement_mt

        if sign == "IN":
            summary["total_in_nsv_bbl"] += movement_nsv
            summary["total_in_lt"] += movement_lt
            summary["total_in_mt"] += movement_mt

        if sign == "OUT":
            summary["total_out_nsv_bbl"] += movement_nsv
            summary["total_out_lt"] += movement_lt
            summary["total_out_mt"] += movement_mt

        summary["closing_nsv_bbl"] = row.running_balance_nsv_bbl or 0
        summary["closing_lt"] = row.running_balance_lt or 0
        summary["closing_mt"] = row.running_balance_mt or 0

    return list(summary_map.values())


@router.get(
    "/tank-stock-ledger/daily-summary",
    response_model=list[TankStockLedgerDailySummaryResponse],
)
def get_tank_stock_ledger_daily_summary(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = "Active",
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Tank Stock Ledger",
        db,
    )

    date_from_value = parse_date_filter(date_from, "Date From")
    date_to_value = parse_date_filter(date_to, "Date To")

    if date_from_value is None or date_to_value is None:
        raise HTTPException(
            status_code=400,
            detail="Date From and Date To are required for daily summary",
        )

    ledger_rows = get_tank_stock_rows_for_daily_summary(
        db=db,
        location_code=location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_to_value=date_to_value,
        status=status,
    )

    return build_tank_stock_daily_summary_rows(
        db=db,
        ledger_rows=ledger_rows,
        date_from_value=date_from_value,
        date_to_value=date_to_value,
    )


@router.get(
    "/out-turn-report",
    response_model=list[OutTurnReportResponse],
)
def get_out_turn_report(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Out-Turn Report",
        db,
    )

    rows = get_out_turn_report_rows_from_transactions(
        db=db,
        location_code=location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_from=date_from,
        date_to=date_to,
    )

    return [
        build_out_turn_report_response_from_transaction(row, db)
        for row in rows
    ]


@router.get("/out-turn-report/validation")
def validate_out_turn_report_tank_sequence(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Out-Turn Report",
        db,
    )

    date_from_value = parse_date_filter(date_from, "Date From")
    date_to_value = parse_date_filter(date_to, "Date To")

    continuity_rows = get_out_turn_report_rows(
        db=db,
        location_code=location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_from=None,
        date_to=date_to,
        status="Active",
    )

    visible_rows = []

    for row in continuity_rows:
        if date_from_value and row.accounting_date and row.accounting_date < date_from_value:
            continue

        if date_to_value and row.accounting_date and row.accounting_date > date_to_value:
            continue

        visible_rows.append(row)

    grouped_rows = {}

    for row in continuity_rows:
        key = (
            row.location_code,
            row.tank_asset_code,
            row.product_name or "",
        )

        if key not in grouped_rows:
            grouped_rows[key] = []

        grouped_rows[key].append(row)

    visible_ledger_ids = {row.id for row in visible_rows}

    issues = []

    for key, group_rows in grouped_rows.items():
        location, tank, product = key

        sorted_group_rows = sorted(
            group_rows,
            key=lambda row: (
                row.accounting_date or date.min,
                get_ledger_operation_datetime(row) or datetime.min,
                row.id,
            ),
        )

        previous_row = None

        for row in sorted_group_rows:
            if previous_row is None:
                expected_previous_nsv = 0
            else:
                previous_snapshot = get_stock_snapshot_values(previous_row)
                expected_previous_nsv = previous_snapshot["nsv"]

            if row.id in visible_ledger_ids:
                actual_previous_nsv = safe_float(row.previous_stock_nsv_bbl)

                if round(actual_previous_nsv, 3) != round(expected_previous_nsv, 3):
                    issues.append(
                        {
                            "ledger_id": row.id,
                            "ticket_number": row.ticket_number,
                            "location_code": location,
                            "tank_asset_code": tank,
                            "product_name": product or None,
                            "expected_previous_nsv_bbl": round(
                                expected_previous_nsv,
                                3,
                            ),
                            "actual_previous_nsv_bbl": round(
                                actual_previous_nsv,
                                3,
                            ),
                            "message": (
                                "Previous stock does not match previous row "
                                "of the same tank/product sequence. Run ledger rebuild."
                            ),
                        }
                    )

            previous_row = row

    return {
        "rows_checked": len(visible_rows),
        "groups_checked": len(grouped_rows),
        "issues_count": len(issues),
        "issues": issues,
    }


@router.get(
    "/material-balance-report",
    response_model=MaterialBalanceDynamicReportResponse,
)
def get_material_balance_report(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = "Active",
    unit: str | None = "nsv",
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Material Balance Report",
        db,
    )

    cleaned_location_code = clean_optional_text(location_code)

    if not cleaned_location_code:
        raise HTTPException(
            status_code=400,
            detail="Location is required for configurable Material Balance Report",
        )

    date_from_value = parse_date_filter(date_from, "Date From")
    date_to_value = parse_date_filter(date_to, "Date To")

    if date_from_value is None or date_to_value is None:
        raise HTTPException(
            status_code=400,
            detail="Date From and Date To are required for Material Balance Report",
        )

    unit_key = normalize_material_balance_code_value(unit).lower()

    if unit_key not in ["gsv", "nsv", "lt", "mt"]:
        raise HTTPException(
            status_code=400,
            detail="Unit must be one of: gsv, nsv, lt, mt",
        )

    template = get_active_material_balance_template_for_location(
        db=db,
        location_code=cleaned_location_code,
    )

    columns = get_active_material_balance_template_columns(
        db=db,
        template_id=template.id,
    )

    ledger_rows = get_material_balance_rows_for_continuity(
        db=db,
        location_code=cleaned_location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_to_value=date_to_value,
        status=status,
    )

    tank_rows = build_dynamic_material_balance_tank_rows(
        db=db,
        ledger_rows=ledger_rows,
        columns=columns,
        date_from_value=date_from_value,
        date_to_value=date_to_value,
        unit_key=unit_key,
    )

    cleaned_tank_asset_code = clean_optional_text(tank_asset_code)

    if cleaned_tank_asset_code:
        report_rows = tank_rows
    else:
        report_rows = consolidate_dynamic_material_balance_rows_by_location(
            tank_rows=tank_rows,
            columns=columns,
        )

    return {
        "template": {
            "id": template.id,
            "location_code": template.location_code,
            "template_name": template.template_name,
        },
        "columns": build_dynamic_material_balance_columns_response(columns),
        "rows": report_rows,
    }


def build_fso_otr_report(
    db: Session,
    location_code: str,
    fso_asset_code: str,
    date_from: date,
    date_to: date,
    shuttle_number: str | None = None,
):
    from app.services.transaction_helpers import approved_transaction_not_on_correction_hold

    loc_code = clean_optional_text(location_code)
    asset_code = clean_optional_text(fso_asset_code)
    sn = clean_optional_text(shuttle_number)

    q = (
        db.query(OperationTransaction, OperationTransactionValue)
        .join(OperationTransactionValue, OperationTransactionValue.transaction_id == OperationTransaction.id)
        .filter(
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
            approved_transaction_not_on_correction_hold(db),
            OperationTransaction.origin_location_code == loc_code,
            OperationTransaction.primary_asset_code == asset_code,
            OperationTransaction.operation_date >= date_from,
            OperationTransaction.operation_date <= date_to,
            OperationTransactionValue.field_code == "fso_payload",
        )
        .order_by(OperationTransaction.operation_date.asc(), OperationTransaction.id.asc())
    )
    if sn:
        q = q.filter(OperationTransaction.convoy_number == sn)

    rows = []
    totals = {
        "receipt": 0.0,
        "export": 0.0,
        "movement": 0.0,
        "variance": 0.0,
        "compare_variance": 0.0,
    }

    for tx, val in q.all():
        payload = val.field_value if isinstance(val.field_value, dict) else {}
        meta = payload.get("meta") or {}
        inputs = payload.get("inputs") or {}
        net = ((payload.get("calculated") or {}).get("net") or {})

        event_time = inputs.get("event_time")
        op_label = str(meta.get("operation_label") or "").strip() or "FSO"
        op_sign = str(meta.get("operation_sign") or "").strip().upper()

        net_stock = float(safe_float(net.get("net_stock_bbl")))
        net_water = float(safe_float(net.get("net_water_bbl")))
        movement_qty = abs(net_stock) + abs(net_water)

        vessel_qty = float(safe_float(inputs.get("vessel_quantity_bbl")))
        variance = abs(net_stock + net_water) - vessel_qty

        src_discharge = float(safe_float(meta.get("source_shuttle_discharge_bbl")))
        compare_var = movement_qty - src_discharge if op_sign == "IN" and src_discharge > 0 else 0.0

        setting = get_active_location_day_setting(db, loc_code, tx.operation_date)
        day_start = setting.day_start_time if setting else datetime_time(0, 0)
        acc_date = compute_accounting_date(tx.operation_date, event_time, day_start)

        row = {
            "transaction_id": tx.id,
            "ticket_number": get_transaction_ticket_number(tx),
            "accounting_date": acc_date,
            "operation_date": tx.operation_date,
            "event_time": event_time,
            "location_code": loc_code,
            "fso_asset_code": asset_code,
            "shuttle_number": inputs.get("shuttle_number") or meta.get("shuttle_number") or tx.convoy_number,
            "operation_label": op_label,
            "operation_sign": op_sign,
            "vessel_name": inputs.get("vessel_name"),
            "vessel_quantity_bbl": vessel_qty,
            "opening_stock_bbl": float(safe_float(inputs.get("opening_stock_bbl"))),
            "opening_water_bbl": float(safe_float(inputs.get("opening_water_bbl"))),
            "closing_stock_bbl": float(safe_float(inputs.get("closing_stock_bbl"))),
            "closing_water_bbl": float(safe_float(inputs.get("closing_water_bbl"))),
            "net_stock_bbl": net_stock,
            "net_water_bbl": net_water,
            "movement_qty_bbl": movement_qty,
            "variance_bbl": variance,
            "source_shuttle_discharge_bbl": src_discharge,
            "compare_variance_bbl": compare_var,
            "remarks": inputs.get("remarks"),
        }
        rows.append(row)

        totals["movement"] += movement_qty
        totals["variance"] += variance
        totals["compare_variance"] += compare_var
        if op_sign == "IN":
            totals["receipt"] += movement_qty
        elif op_sign == "OUT":
            totals["export"] += movement_qty

    return rows, totals


def build_fso_material_balance(
    db: Session,
    location_code: str,
    fso_asset_code: str,
    date_from: date,
    date_to: date,
):
    from app.services.transaction_helpers import approved_transaction_not_on_correction_hold

    loc_code = clean_optional_text(location_code)
    asset_code = clean_optional_text(fso_asset_code)

    q = (
        db.query(OperationTransaction, OperationTransactionValue)
        .join(OperationTransactionValue, OperationTransactionValue.transaction_id == OperationTransaction.id)
        .filter(
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
            approved_transaction_not_on_correction_hold(db),
            OperationTransaction.origin_location_code == loc_code,
            OperationTransaction.primary_asset_code == asset_code,
            OperationTransaction.operation_date >= date_from,
            OperationTransaction.operation_date <= date_to,
            OperationTransactionValue.field_code == "fso_payload",
        )
        .order_by(OperationTransaction.operation_date.asc(), OperationTransaction.id.asc())
    )

    buckets = {}
    for tx, val in q.all():
        payload = val.field_value if isinstance(val.field_value, dict) else {}
        meta = payload.get("meta") or {}
        inputs = payload.get("inputs") or {}
        net = ((payload.get("calculated") or {}).get("net") or {})

        setting = get_active_location_day_setting(db, loc_code, tx.operation_date)
        day_start = setting.day_start_time if setting else datetime_time(0, 0)
        acc_date = compute_accounting_date(tx.operation_date, inputs.get("event_time"), day_start)

        buckets.setdefault(acc_date, []).append((tx, meta, inputs, net))

    dates = sorted([d for d in buckets.keys() if d >= date_from and d <= date_to])
    rows = []
    prev_physical_close = None

    for acc_date in dates:
        items = buckets[acc_date]

        def sort_key(item):
            tx, meta, inputs, net = item
            t = inputs.get("event_time") or "00:00"
            return (str(tx.operation_date), str(t), tx.id)

        items = sorted(items, key=sort_key)

        if prev_physical_close is None:
            opening = float(safe_float(items[0][2].get("opening_stock_bbl")))
        else:
            opening = prev_physical_close

        receipt = 0.0
        export = 0.0

        for tx, meta, inputs, net in items:
            op_sign = str(meta.get("operation_sign") or "").strip().upper()
            net_stock = float(safe_float(net.get("net_stock_bbl")))
            net_water = float(safe_float(net.get("net_water_bbl")))
            qty = abs(net_stock) + abs(net_water)

            if op_sign == "IN":
                receipt += qty
            elif op_sign == "OUT":
                export += qty

        book_close = opening + receipt - export

        last_inputs = items[-1][2]
        physical_close = float(safe_float(last_inputs.get("closing_stock_bbl")))
        physical_close_water = float(safe_float(last_inputs.get("closing_water_bbl")))
        loss_gain = physical_close - book_close

        rows.append(
            {
                "accounting_date": acc_date,
                "opening_stock_bbl": opening,
                "receipt_bbl": receipt,
                "export_bbl": export,
                "book_closing_bbl": book_close,
                "physical_closing_bbl": physical_close,
                "physical_closing_water_bbl": physical_close_water,
                "loss_gain_bbl": loss_gain,
            }
        )

        prev_physical_close = physical_close

    return rows


def build_fso_outturn_report(
    db: Session,
    location_code: str,
    fso_asset_code: str,
    date_from: date,
    date_to: date,
):
    from app.services.transaction_helpers import approved_transaction_not_on_correction_hold

    loc_code = clean_optional_text(location_code)
    asset_code = clean_optional_text(fso_asset_code)

    q = (
        db.query(OperationTransaction, OperationTransactionValue)
        .join(OperationTransactionValue, OperationTransactionValue.transaction_id == OperationTransaction.id)
        .filter(
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
            approved_transaction_not_on_correction_hold(db),
            OperationTransaction.origin_location_code == loc_code,
            OperationTransaction.primary_asset_code == asset_code,
            OperationTransaction.operation_date >= date_from,
            OperationTransaction.operation_date <= date_to,
            OperationTransactionValue.field_code == "fso_payload",
        )
        .order_by(OperationTransaction.operation_date.asc(), OperationTransaction.id.asc())
    )

    def _sf(v):
        try:
            return float(v or 0.0)
        except Exception:
            return 0.0

    def _abs_qty(net_stock, net_water):
        return abs(_sf(net_stock)) + abs(_sf(net_water))

    buckets = {}
    for tx, val in q.all():
        payload = val.field_value if isinstance(val.field_value, dict) else {}
        meta = payload.get("meta") or {}
        inputs = payload.get("inputs") or {}
        net = ((payload.get("calculated") or {}).get("net") or {})

        event_time = inputs.get("event_time")

        setting = get_active_location_day_setting(db, loc_code, tx.operation_date)
        day_start = setting.day_start_time if setting else datetime_time(0, 0)
        acc_date = compute_accounting_date(tx.operation_date, event_time, day_start)

        shuttle_no = (
            inputs.get("shuttle_number")
            or meta.get("shuttle_number")
            or tx.convoy_number
            or ""
        ).strip()
        if shuttle_no == "":
            continue

        key = (acc_date, shuttle_no)
        buckets.setdefault(key, {"receipt": 0.0, "discharge": 0.0})

        op_sign = str(meta.get("operation_sign") or "").strip().upper()

        net_stock = net.get("net_stock_bbl")
        net_water = net.get("net_water_bbl")
        qty = _abs_qty(net_stock, net_water)

        if op_sign == "IN":
            buckets[key]["receipt"] += qty

        src = meta.get("source_shuttle_discharge_bbl")
        if src is not None:
            buckets[key]["discharge"] = max(buckets[key]["discharge"], _sf(src))

    rows = []
    totals = {"discharge": 0.0, "receipt": 0.0, "variance": 0.0}

    for (acc_date, shuttle_no) in sorted(buckets.keys()):
        discharge = float(buckets[(acc_date, shuttle_no)]["discharge"])
        receipt = float(buckets[(acc_date, shuttle_no)]["receipt"])
        variance = receipt - discharge
        pct = (variance / discharge * 100.0) if discharge != 0 else 0.0

        rows.append(
            {
                "accounting_date": acc_date,
                "shuttle_number": shuttle_no,
                "shuttle_discharge_bbl": discharge,
                "fso_receipt_bbl": receipt,
                "variance_bbl": variance,
                "variance_pct": pct,
            }
        )

        totals["discharge"] += discharge
        totals["receipt"] += receipt
        totals["variance"] += variance

    totals_pct = (totals["variance"] / totals["discharge"] * 100.0) if totals["discharge"] != 0 else 0.0
    return rows, totals, totals_pct


def _xlsx_autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)


@router.get("/fso/otr", response_model=FSOOTRReportResponse)
def get_fso_otr_report(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    shuttle_number: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")

    rows, totals = build_fso_otr_report(db, location_code, fso_asset_code, df, dt, shuttle_number)
    return {
        "rows": rows,
        "total_receipt_bbl": totals["receipt"],
        "total_export_bbl": totals["export"],
        "total_movement_bbl": totals["movement"],
        "total_variance_bbl": totals["variance"],
        "total_compare_variance_bbl": totals["compare_variance"],
    }


@router.get("/fso/material-balance", response_model=FSOMaterialBalanceReportResponse)
def get_fso_material_balance_report(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")
    rows = build_fso_material_balance(db, location_code, fso_asset_code, df, dt)
    return {"rows": rows}


@router.get("/fso/outturn", response_model=FSOOutturnReportResponse)
def fso_report_outturn(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")

    rows, totals, totals_pct = build_fso_outturn_report(db, location_code, fso_asset_code, df, dt)
    return {
        "rows": rows,
        "total_shuttle_discharge_bbl": totals["discharge"],
        "total_fso_receipt_bbl": totals["receipt"],
        "total_variance_bbl": totals["variance"],
        "total_variance_pct": totals_pct,
    }


@router.get("/fso/otr/export/xlsx")
def fso_report_otr_xlsx(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    shuttle_number: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")
    rows, totals = build_fso_otr_report(db, location_code, fso_asset_code, df, dt, shuttle_number)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FSO OTR"

    headers = [
        "Ticket", "Acc Date", "Op Date", "Time", "Operation", "Sign",
        "Shuttle", "Vessel", "Vessel Qty",
        "Open Stock", "Open Water", "Close Stock", "Close Water",
        "Net Stock", "Net Water", "Move Qty",
        "Variance", "Shuttle Discharge", "Compare Var", "Remarks",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            r["ticket_number"],
            str(r["accounting_date"]),
            str(r["operation_date"]),
            r.get("event_time") or "",
            r["operation_label"],
            r["operation_sign"],
            r.get("shuttle_number") or "",
            r.get("vessel_name") or "",
            round(float(r["vessel_quantity_bbl"]), 3),
            round(float(r["opening_stock_bbl"]), 3),
            round(float(r["opening_water_bbl"]), 3),
            round(float(r["closing_stock_bbl"]), 3),
            round(float(r["closing_water_bbl"]), 3),
            round(float(r["net_stock_bbl"]), 3),
            round(float(r["net_water_bbl"]), 3),
            round(float(r["movement_qty_bbl"]), 3),
            round(float(r["variance_bbl"]), 3),
            round(float(r["source_shuttle_discharge_bbl"]), 3),
            round(float(r["compare_variance_bbl"]), 3),
            r.get("remarks") or "",
        ])

    ws2 = wb.create_sheet("Totals")
    ws2.append(["Total Receipt", totals["receipt"]])
    ws2.append(["Total Export", totals["export"]])
    ws2.append(["Total Movement", totals["movement"]])
    ws2.append(["Total Variance", totals["variance"]])
    ws2.append(["Total Compare Variance", totals["compare_variance"]])

    _xlsx_autofit(ws)
    _xlsx_autofit(ws2)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"fso_otr_{location_code}_{fso_asset_code}_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/fso/outturn/export/xlsx")
def fso_report_outturn_xlsx(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")

    rows, totals, totals_pct = build_fso_outturn_report(db, location_code, fso_asset_code, df, dt)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FSO Outturn"

    headers = ["Acc Date", "Shuttle Number", "Shuttle Discharge", "FSO Receipt", "Variance", "Variance %"]
    ws.append(headers)

    for r in rows:
        ws.append([
            str(r["accounting_date"]),
            r["shuttle_number"],
            round(float(r["shuttle_discharge_bbl"]), 3),
            round(float(r["fso_receipt_bbl"]), 3),
            round(float(r["variance_bbl"]), 3),
            round(float(r["variance_pct"]), 3),
        ])

    ws2 = wb.create_sheet("Totals")
    ws2.append(["Total Shuttle Discharge", totals["discharge"]])
    ws2.append(["Total FSO Receipt", totals["receipt"]])
    ws2.append(["Total Variance", totals["variance"]])
    ws2.append(["Total Variance %", totals_pct])

    _xlsx_autofit(ws)
    _xlsx_autofit(ws2)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"fso_outturn_{location_code}_{fso_asset_code}_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/fso/material-balance/export/xlsx")
def fso_report_mb_xlsx(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")
    rows = build_fso_material_balance(db, location_code, fso_asset_code, df, dt)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FSO Material Balance"

    headers = [
        "Acc Date", "Opening", "Receipt", "Export",
        "Book Closing", "Physical Closing", "Closing Water", "Loss/Gain",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            str(r["accounting_date"]),
            round(float(r["opening_stock_bbl"]), 3),
            round(float(r["receipt_bbl"]), 3),
            round(float(r["export_bbl"]), 3),
            round(float(r["book_closing_bbl"]), 3),
            round(float(r["physical_closing_bbl"]), 3),
            round(float(r["physical_closing_water_bbl"]), 3),
            round(float(r["loss_gain_bbl"]), 3),
        ])

    _xlsx_autofit(ws)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"fso_mb_{location_code}_{fso_asset_code}_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/tank-stock-ledger/rebuild")
def rebuild_tank_stock_ledger(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Tank Stock Ledger",
        db,
    )

    query = db.query(TankStockLedger).filter(
        TankStockLedger.status == "Active",
    )

    cleaned_location_code = clean_optional_text(location_code)
    cleaned_tank_asset_code = clean_optional_text(tank_asset_code)
    cleaned_product_name = clean_optional_text(product_name)

    if cleaned_location_code:
        query = query.filter(TankStockLedger.location_code.ilike(cleaned_location_code))

    if cleaned_tank_asset_code:
        query = query.filter(TankStockLedger.tank_asset_code.ilike(cleaned_tank_asset_code))

    if cleaned_product_name:
        query = query.filter(TankStockLedger.product_name.ilike(cleaned_product_name))

    rows = query.all()

    group_keys = set()

    for row in rows:
        group_keys.add(
            (
                row.location_code,
                row.tank_asset_code,
                row.product_name,
            )
        )

    for location, tank_asset, product in group_keys:
        rebuild_tank_stock_running_balances(
            db=db,
            location_code=location,
            tank_asset_code=tank_asset,
            product_name=product,
        )

    create_audit_log(
        db=db,
        module_name="Tank Stock Ledger",
        action="Rebuild Tank Stock Ledger",
        current_user=current_user,
        entity_type="TankStockLedger",
        entity_id=None,
        entity_label="Tank Stock Ledger Rebuild",
        remarks="Rebuilt stock movements from chronological tank stock snapshots",
        request_path="/tank-stock-ledger/rebuild",
        details={
            "location_code": cleaned_location_code,
            "tank_asset_code": cleaned_tank_asset_code,
            "product_name": cleaned_product_name,
            "groups_rebuilt": len(group_keys),
            "rows_scanned": len(rows),
        },
    )

    db.commit()

    return {
        "message": "Tank Stock Ledger rebuilt successfully",
        "groups_rebuilt": len(group_keys),
        "rows_scanned": len(rows),
    }


def normalize_jsonb_value(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.isoformat()

    if hasattr(value, "isoformat"):
        return value.isoformat()

    if isinstance(value, list):
        return [
            normalize_jsonb_value(item)
            for item in value
        ]

    if isinstance(value, dict):
        return {
            str(key): normalize_jsonb_value(item_value)
            for key, item_value in value.items()
        }

    return value


def get_tank_gauging_payload_for_transaction(
    db: Session,
    transaction_id: int,
):
    payload_row = (
        db.query(OperationTransactionValue)
        .filter(
            OperationTransactionValue.transaction_id == transaction_id,
            OperationTransactionValue.field_code == "tank_gauging_payload",
        )
        .first()
    )

    if payload_row is None or payload_row.field_value is None:
        return None

    if not isinstance(payload_row.field_value, dict):
        return None

    return payload_row.field_value


def parse_payload_gauging_datetime(payload: dict):
    inputs = payload.get("inputs") or {}

    gauging_date = clean_optional_text(inputs.get("gaugingDate"))
    gauging_time = clean_optional_text(inputs.get("gaugingTime"))

    if not gauging_date or not gauging_time:
        return None

    try:
        return datetime.fromisoformat(f"{gauging_date}T{gauging_time}")
    except ValueError:
        return None


def resolve_transaction_datetime_for_accounting_day(
    transaction: OperationTransaction,
    payload: dict,
):
    if transaction.operation_start_datetime is not None:
        return transaction.operation_start_datetime

    payload_datetime = parse_payload_gauging_datetime(payload)

    if payload_datetime is not None:
        return payload_datetime

    raise HTTPException(
        status_code=400,
        detail=(
            "Operation Start Date/Time or Tank Gauging Date/Time is required "
            "to calculate the Location Accounting Day."
        ),
    )


def calculate_accounting_window_from_setting(
    setting: LocationAccountingDaySetting,
    transaction_datetime: datetime,
):
    transaction_date = transaction_datetime.date()
    transaction_time = transaction_datetime.time()

    start_time = setting.day_start_time
    end_time = setting.day_end_time

    is_overnight_window = end_time < start_time

    if is_overnight_window:
        if transaction_time >= start_time:
            accounting_date = transaction_date
        else:
            accounting_date = transaction_date - timedelta(days=1)

        accounting_day_start = datetime.combine(accounting_date, start_time)
        accounting_day_end = datetime.combine(
            accounting_date + timedelta(days=1),
            end_time,
        )

    else:
        if transaction_time >= start_time:
            accounting_date = transaction_date
        else:
            accounting_date = transaction_date - timedelta(days=1)

        accounting_day_start = datetime.combine(accounting_date, start_time)
        accounting_day_end = datetime.combine(accounting_date, end_time)

    return {
        "accounting_date": accounting_date,
        "accounting_day_start": accounting_day_start,
        "accounting_day_end": accounting_day_end,
    }


def get_location_accounting_day_for_transaction(
    db: Session,
    location_code: str,
    transaction_datetime: datetime,
):
    cleaned_location_code = clean_optional_text(location_code)

    if not cleaned_location_code:
        raise HTTPException(
            status_code=400,
            detail="Location is required to calculate accounting day",
        )

    active_settings = (
        db.query(LocationAccountingDaySetting)
        .filter(
            LocationAccountingDaySetting.location_code.ilike(cleaned_location_code),
            LocationAccountingDaySetting.status == "Active",
        )
        .order_by(
            LocationAccountingDaySetting.effective_from.desc(),
            LocationAccountingDaySetting.id.desc(),
        )
        .all()
    )

    if len(active_settings) == 0:
        raise HTTPException(
            status_code=400,
            detail=(
                "No Active Location Accounting Day Setting found for "
                f"{cleaned_location_code}. Configure it before approving "
                "Tank Gauging tickets."
            ),
        )

    matching_options = []

    for setting in active_settings:
        window = calculate_accounting_window_from_setting(
            setting=setting,
            transaction_datetime=transaction_datetime,
        )

        accounting_date = window["accounting_date"]

        effective_to = setting.effective_to or date(9999, 12, 31)

        if setting.effective_from <= accounting_date <= effective_to:
            if (
                window["accounting_day_start"]
                <= transaction_datetime
                <= window["accounting_day_end"]
            ):
                matching_options.append(
                    {
                        "setting": setting,
                        "window": window,
                    }
                )

    if len(matching_options) == 0:
        raise HTTPException(
            status_code=400,
            detail=(
                "No effective Location Accounting Day Setting matched this "
                "transaction date/time. Check Effective From/To settings."
            ),
        )

    selected = matching_options[0]
    selected_setting = selected["setting"]
    selected_window = selected["window"]

    return {
        "setting_id": selected_setting.id,
        "accounting_date": selected_window["accounting_date"],
        "accounting_day_start": selected_window["accounting_day_start"],
        "accounting_day_end": selected_window["accounting_day_end"],
    }


def get_ledger_sort_datetime(ledger: TankStockLedger):
    if ledger.accounting_day_start is not None:
        return ledger.accounting_day_start

    if ledger.operation_date is not None:
        return datetime.combine(ledger.operation_date, datetime_time(0, 0))

    return datetime.min


def get_previous_active_ledger_row(
    db: Session,
    location_code: str,
    tank_asset_code: str,
    product_name: str | None,
    transaction_datetime: datetime,
    exclude_ledger_id: int | None = None,
):
    query = db.query(TankStockLedger).filter(
        TankStockLedger.status == "Active",
        TankStockLedger.location_code.ilike(location_code),
        TankStockLedger.tank_asset_code.ilike(tank_asset_code),
    )

    cleaned_product_name = clean_optional_text(product_name)

    if cleaned_product_name:
        query = query.filter(TankStockLedger.product_name.ilike(cleaned_product_name))
    else:
        query = query.filter(TankStockLedger.product_name == None)

    if exclude_ledger_id is not None:
        query = query.filter(TankStockLedger.id != exclude_ledger_id)

    candidate_rows = query.all()

    previous_rows = []

    for row in candidate_rows:
        row_datetime = row.accounting_day_start

        try:
            payload = row.source_payload or {}
            payload_inputs = payload.get("inputs") or {}
            gauging_date = clean_optional_text(payload_inputs.get("gaugingDate"))
            gauging_time = clean_optional_text(payload_inputs.get("gaugingTime"))

            if gauging_date and gauging_time:
                row_datetime = datetime.fromisoformat(
                    f"{gauging_date}T{gauging_time}"
                )
        except Exception:
            row_datetime = None

        if row_datetime is None:
            row_datetime = datetime.combine(row.operation_date, datetime_time(0, 0))

        if row_datetime < transaction_datetime:
            previous_rows.append((row_datetime, row.id, row))

    if not previous_rows:
        return None

    previous_rows.sort(key=lambda item: (item[0], item[1]))

    return previous_rows[-1][2]


def calculate_stock_movement_from_snapshot(
    operation_sign: str,
    current_gsv_bbl: float,
    current_nsv_bbl: float,
    current_lt: float,
    current_mt: float,
    previous_ledger: TankStockLedger | None,
):
    sign = str(operation_sign or "").upper()

    previous_gsv_bbl = 0
    previous_nsv_bbl = 0
    previous_lt = 0
    previous_mt = 0

    if previous_ledger is not None:
        previous_gsv_bbl = safe_float(
            previous_ledger.stock_gsv_bbl
            if previous_ledger.stock_gsv_bbl is not None
            else previous_ledger.running_balance_gsv_bbl
        )
        previous_nsv_bbl = safe_float(
            previous_ledger.stock_nsv_bbl
            if previous_ledger.stock_nsv_bbl is not None
            else previous_ledger.running_balance_nsv_bbl
        )
        previous_lt = safe_float(
            previous_ledger.stock_lt
            if previous_ledger.stock_lt is not None
            else previous_ledger.running_balance_lt
        )
        previous_mt = safe_float(
            previous_ledger.stock_mt
            if previous_ledger.stock_mt is not None
            else previous_ledger.running_balance_mt
        )

    if sign == "SET":
        movement_gsv_bbl = current_gsv_bbl
        movement_nsv_bbl = current_nsv_bbl
        movement_lt = current_lt
        movement_mt = current_mt

    elif sign == "IN":
        movement_gsv_bbl = max(current_gsv_bbl - previous_gsv_bbl, 0)
        movement_nsv_bbl = max(current_nsv_bbl - previous_nsv_bbl, 0)
        movement_lt = max(current_lt - previous_lt, 0)
        movement_mt = max(current_mt - previous_mt, 0)

    elif sign == "OUT":
        movement_gsv_bbl = max(previous_gsv_bbl - current_gsv_bbl, 0)
        movement_nsv_bbl = max(previous_nsv_bbl - current_nsv_bbl, 0)
        movement_lt = max(previous_lt - current_lt, 0)
        movement_mt = max(previous_mt - current_mt, 0)

    elif sign == "NEUTRAL":
        movement_gsv_bbl = 0
        movement_nsv_bbl = 0
        movement_lt = 0
        movement_mt = 0

    else:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid Tank Operation Sign: {operation_sign}",
        )

    return {
        "previous_gsv_bbl": previous_gsv_bbl,
        "previous_nsv_bbl": previous_nsv_bbl,
        "previous_lt": previous_lt,
        "previous_mt": previous_mt,
        "movement_gsv_bbl": movement_gsv_bbl,
        "movement_nsv_bbl": movement_nsv_bbl,
        "movement_lt": movement_lt,
        "movement_mt": movement_mt,
    }


def is_tank_gauging_transaction(
    db: Session,
    transaction: OperationTransaction,
):
    if transaction.operation_template_id is None:
        return False

    template = (
        db.query(OperationTemplate)
        .filter(OperationTemplate.id == transaction.operation_template_id)
        .first()
    )

    if not template:
        return False

    entry_layout_type = str(template.entry_layout_type or "").strip()
    calculation_engine = str(template.calculation_engine or "").strip()

    if entry_layout_type == "Tank Gauging":
        return True

    if calculation_engine == "Tank Quantity":
        return True

    payload = get_tank_gauging_payload_for_transaction(
        db=db,
        transaction_id=transaction.id,
    )

    return payload is not None


def rebuild_tank_stock_running_balances(
    db: Session,
    location_code: str,
    tank_asset_code: str,
    product_name: str | None,
):
    query = db.query(TankStockLedger).filter(
        TankStockLedger.location_code.ilike(location_code),
        TankStockLedger.tank_asset_code.ilike(tank_asset_code),
        TankStockLedger.status == "Active",
    )

    cleaned_product_name = clean_optional_text(product_name)

    if cleaned_product_name:
        query = query.filter(
            TankStockLedger.product_name.ilike(cleaned_product_name)
        )
    else:
        query = query.filter(TankStockLedger.product_name == None)

    ledger_rows = query.all()

    sortable_rows = []

    for row in ledger_rows:
        row_datetime = row.accounting_day_start

        try:
            payload = row.source_payload or {}
            payload_inputs = payload.get("inputs") or {}
            gauging_date = clean_optional_text(payload_inputs.get("gaugingDate"))
            gauging_time = clean_optional_text(payload_inputs.get("gaugingTime"))

            if gauging_date and gauging_time:
                row_datetime = datetime.fromisoformat(
                    f"{gauging_date}T{gauging_time}"
                )
        except Exception:
            row_datetime = None

        if row_datetime is None:
            row_datetime = datetime.combine(row.operation_date, datetime_time(0, 0))

        sortable_rows.append((row_datetime, row.id, row))

    sortable_rows.sort(key=lambda item: (item[0], item[1]))

    previous_row = None

    for row_datetime, _row_id, row in sortable_rows:
        if (
            row.accounting_date is None
            or row.accounting_day_start is None
            or row.accounting_day_end is None
            or row.accounting_day_setting_id is None
        ):
            try:
                payload = row.source_payload or {}

                transaction_datetime = resolve_transaction_datetime_for_accounting_day(
                    transaction=db.query(OperationTransaction)
                    .filter(OperationTransaction.id == row.transaction_id)
                    .first(),
                    payload=payload,
                )

                accounting_day = get_location_accounting_day_for_transaction(
                    db=db,
                    location_code=row.location_code,
                    transaction_datetime=transaction_datetime,
                )

                row.accounting_date = accounting_day["accounting_date"]
                row.accounting_day_start = accounting_day["accounting_day_start"]
                row.accounting_day_end = accounting_day["accounting_day_end"]
                row.accounting_day_setting_id = accounting_day["setting_id"]

            except Exception:
                pass

        current_gsv_bbl = safe_float(row.stock_gsv_bbl)
        current_nsv_bbl = safe_float(row.stock_nsv_bbl)
        current_lt = safe_float(row.stock_lt)
        current_mt = safe_float(row.stock_mt)

        if current_gsv_bbl == 0 and current_nsv_bbl == 0:
            current_gsv_bbl = safe_float(row.running_balance_gsv_bbl)
            current_nsv_bbl = safe_float(row.running_balance_nsv_bbl)
            current_lt = safe_float(row.running_balance_lt)
            current_mt = safe_float(row.running_balance_mt)

            row.stock_gsv_bbl = current_gsv_bbl
            row.stock_nsv_bbl = current_nsv_bbl
            row.stock_lt = current_lt
            row.stock_mt = current_mt

        movement = calculate_stock_movement_from_snapshot(
            operation_sign=row.tank_operation_sign,
            current_gsv_bbl=current_gsv_bbl,
            current_nsv_bbl=current_nsv_bbl,
            current_lt=current_lt,
            current_mt=current_mt,
            previous_ledger=previous_row,
        )

        row.previous_stock_gsv_bbl = movement["previous_gsv_bbl"]
        row.previous_stock_nsv_bbl = movement["previous_nsv_bbl"]
        row.previous_stock_lt = movement["previous_lt"]
        row.previous_stock_mt = movement["previous_mt"]

        row.movement_gsv_bbl = movement["movement_gsv_bbl"]
        row.movement_nsv_bbl = movement["movement_nsv_bbl"]
        row.movement_lt = movement["movement_lt"]
        row.movement_mt = movement["movement_mt"]

        row.running_balance_gsv_bbl = current_gsv_bbl
        row.running_balance_nsv_bbl = current_nsv_bbl
        row.running_balance_lt = current_lt
        row.running_balance_mt = current_mt

        row.updated_at = datetime.now()

        previous_row = row

    db.flush()


def create_tank_stock_ledger_from_approved_transaction(
    db: Session,
    transaction: OperationTransaction,
    current_user: User,
):
    if transaction.status != "Approved":
        return None

    if not is_tank_gauging_transaction(db, transaction):
        return None

    existing_ledger = (
        db.query(TankStockLedger)
        .filter(
            TankStockLedger.transaction_id == transaction.id,
            TankStockLedger.status.in_(["Active", "Correction Hold"]),
        )
        .first()
    )

    if existing_ledger:
        return existing_ledger

    payload = get_tank_gauging_payload_for_transaction(
        db=db,
        transaction_id=transaction.id,
    )

    if payload is None:
        raise HTTPException(
            status_code=400,
            detail=(
                "Tank Gauging payload is missing. Open Operation Entry, "
                "save the tank gauging ticket, then approve again."
            ),
        )

    inputs = payload.get("inputs") or {}
    calculated = payload.get("calculated") or {}
    payload_asset = payload.get("asset") or {}

    transaction_datetime = resolve_transaction_datetime_for_accounting_day(
        transaction=transaction,
        payload=payload,
    )

    accounting_day = get_location_accounting_day_for_transaction(
        db=db,
        location_code=transaction.origin_location_code,
        transaction_datetime=transaction_datetime,
    )

    tank_operation_code = clean_optional_text(
        inputs.get("tankOperationCode")
    )
    tank_operation_label = clean_optional_text(
        inputs.get("tankOperationLabel")
    )
    tank_operation_category = clean_optional_text(
        inputs.get("tankOperationCategory")
    )
    tank_operation_sign = clean_optional_text(
        inputs.get("tankOperationSign")
    )

    if not tank_operation_code:
        raise HTTPException(
            status_code=400,
            detail=(
                "Tank Operation is missing in Tank Gauging payload. "
                "Open the ticket, select Tank Operation, save, then approve."
            ),
        )

    if not tank_operation_label:
        raise HTTPException(
            status_code=400,
            detail="Tank Operation Label is missing in Tank Gauging payload.",
        )

    if not tank_operation_category:
        raise HTTPException(
            status_code=400,
            detail="Tank Operation Category is missing in Tank Gauging payload.",
        )

    if not tank_operation_sign:
        raise HTTPException(
            status_code=400,
            detail="Tank Operation Sign is missing in Tank Gauging payload.",
        )

    current_stock_gsv_bbl = safe_float(calculated.get("gsvBbl"))
    current_stock_nsv_bbl = safe_float(calculated.get("nsvBbl"))
    current_stock_lt = safe_float(calculated.get("lt"))
    current_stock_mt = safe_float(calculated.get("mt"))

    if current_stock_nsv_bbl == 0 and current_stock_gsv_bbl == 0:
        raise HTTPException(
            status_code=400,
            detail=(
                "Calculated tank quantity is missing or zero. "
                "Open the ticket, verify Tank Gauging calculations, save, then approve."
            ),
        )

    tank_asset = get_asset_by_code(transaction.primary_asset_code, db)

    tank_asset_name = ""

    if tank_asset:
        tank_asset_name = tank_asset.asset_name
    else:
        tank_asset_name = clean_optional_text(
            payload_asset.get("asset_name")
        ) or ""

    created_by_display = get_current_user_display_name(current_user)

    new_ledger = TankStockLedger(
        transaction_id=transaction.id,
        ticket_number=get_transaction_ticket_number(transaction),
        operation_number=transaction.operation_number,
        location_code=transaction.origin_location_code,
        tank_asset_code=transaction.primary_asset_code,
        tank_asset_name=tank_asset_name,
        operation_date=transaction.operation_date,
        product_name=clean_optional_text(transaction.product_name),
        accounting_date=accounting_day["accounting_date"],
        accounting_day_start=accounting_day["accounting_day_start"],
        accounting_day_end=accounting_day["accounting_day_end"],
        accounting_day_setting_id=accounting_day["setting_id"],
        tank_operation_code=tank_operation_code,
        tank_operation_label=tank_operation_label,
        tank_operation_category=tank_operation_category,
        tank_operation_sign=tank_operation_sign,
        movement_gsv_bbl=0,
        movement_nsv_bbl=0,
        movement_lt=0,
        movement_mt=0,
        stock_gsv_bbl=current_stock_gsv_bbl,
        stock_nsv_bbl=current_stock_nsv_bbl,
        stock_lt=current_stock_lt,
        stock_mt=current_stock_mt,
        previous_stock_gsv_bbl=0,
        previous_stock_nsv_bbl=0,
        previous_stock_lt=0,
        previous_stock_mt=0,
        running_balance_gsv_bbl=current_stock_gsv_bbl,
        running_balance_nsv_bbl=current_stock_nsv_bbl,
        running_balance_lt=current_stock_lt,
        running_balance_mt=current_stock_mt,
        source_payload=normalize_jsonb_value(payload),
        status="Active",
        created_by=created_by_display,
        remarks="Auto-created when Tank Gauging ticket was approved",
    )

    db.add(new_ledger)
    db.flush()

    rebuild_tank_stock_running_balances(
        db=db,
        location_code=new_ledger.location_code,
        tank_asset_code=new_ledger.tank_asset_code,
        product_name=new_ledger.product_name,
    )

    db.flush()

    create_audit_log(
        db=db,
        module_name="Tank Stock Ledger",
        action="Create Tank Stock Ledger Entry",
        current_user=current_user,
        entity_type="TankStockLedger",
        entity_id=new_ledger.id,
        entity_label=(
            f"{new_ledger.ticket_number} | "
            f"{new_ledger.tank_asset_code} | "
            f"{new_ledger.tank_operation_label}"
        ),
        ticket_number=new_ledger.ticket_number,
        operation_number=new_ledger.operation_number,
        remarks="Auto-created on Tank Gauging approval",
        request_path="/operation-transactions/{id}/status",
        details={
            "transaction_id": transaction.id,
            "location_code": new_ledger.location_code,
            "tank_asset_code": new_ledger.tank_asset_code,
            "operation_date": str(new_ledger.operation_date),
            "transaction_datetime": transaction_datetime.isoformat(),
            "accounting_date": str(new_ledger.accounting_date),
            "accounting_day_start": (
                new_ledger.accounting_day_start.isoformat()
                if new_ledger.accounting_day_start
                else None
            ),
            "accounting_day_end": (
                new_ledger.accounting_day_end.isoformat()
                if new_ledger.accounting_day_end
                else None
            ),
            "accounting_day_setting_id": new_ledger.accounting_day_setting_id,
            "product_name": new_ledger.product_name,
            "tank_operation_code": new_ledger.tank_operation_code,
            "tank_operation_label": new_ledger.tank_operation_label,
            "tank_operation_category": new_ledger.tank_operation_category,
            "tank_operation_sign": new_ledger.tank_operation_sign,
            "stock_gsv_bbl": new_ledger.stock_gsv_bbl,
            "stock_nsv_bbl": new_ledger.stock_nsv_bbl,
            "stock_lt": new_ledger.stock_lt,
            "stock_mt": new_ledger.stock_mt,
            "previous_stock_gsv_bbl": new_ledger.previous_stock_gsv_bbl,
            "previous_stock_nsv_bbl": new_ledger.previous_stock_nsv_bbl,
            "previous_stock_lt": new_ledger.previous_stock_lt,
            "previous_stock_mt": new_ledger.previous_stock_mt,
            "movement_gsv_bbl": new_ledger.movement_gsv_bbl,
            "movement_nsv_bbl": new_ledger.movement_nsv_bbl,
            "movement_lt": new_ledger.movement_lt,
            "movement_mt": new_ledger.movement_mt,
            "running_balance_gsv_bbl": new_ledger.running_balance_gsv_bbl,
            "running_balance_nsv_bbl": new_ledger.running_balance_nsv_bbl,
            "running_balance_lt": new_ledger.running_balance_lt,
            "running_balance_mt": new_ledger.running_balance_mt,
        },
    )

    return new_ledger
